import argparse
import json
import os
import signal
import time
import atexit
import uuid
import base64
from io import BytesIO
from datetime import datetime
from types import SimpleNamespace

from dotenv import load_dotenv
from flask import Flask, Response, render_template, request, redirect, url_for, send_from_directory, send_file
from flask_cors import CORS
from waitress import serve
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from lib import FREngine, VideoPlayer, StreamState
from lib.utils import list_camera_devices, log_info, init_logger

import warnings
warnings.filterwarnings("ignore")

load_dotenv()
app = Flask(__name__)
CORS(app) 

init_logger()
log_info("Starting FR Session")

videoplayer = VideoPlayer(
    width=int(os.getenv("WIDTH", "1920")),
    height=int(os.getenv("HEIGHT", "1080")),
    fps=int(os.getenv("FPS", "25")),
    jpg_quality=int(os.getenv("STREAM_JPG_QUALITY", "75"))
)

fr_instance = FREngine(
    videoplayer=videoplayer,
    inference_width=int(os.getenv("INFERENCE_WIDTH", "640")),
    inference_height=int(os.getenv("INFERENCE_HEIGHT", "480")),
)

# Config from environment 
config = SimpleNamespace(
    ip=os.getenv("APP_IP", "0.0.0.0"),
    port=int(os.getenv("APP_PORT", "1334")),
    env=os.getenv("APP_ENV", "production").lower(), # "production" or "development"
)

# Ensure fr_instance runs .cleanup() before process termination
atexit.register(lambda sig, frame: fr_instance.cleanup(force_exit=True))

# ───────────────────────────── API Routes ─────────────────────────────────

@app.route("/api/start_stream", methods=["POST"])
def start_stream():
    """API for frontend to start videoplayer stream"""
    payload = request.get_json(silent=True)
    stream_src = payload.get("stream_src")

    if videoplayer.is_started():
        return _json({"stream": False, "message": "Stream already started!"})

    videoplayer.start_stream(stream_src)

    # Poll the stream thread until it starts (if not assume it failed)
    for delay in [0, 0.1, 0.2, 0.5, 1]:
        time.sleep(delay)
        if videoplayer.streamThread and videoplayer.streamThread.is_alive():
            break
        log_info(f"Stream thread not started, checking again in {delay}s...")

    if not videoplayer.streamThread or not videoplayer.streamThread.is_alive():
        log_info("Stream thread couldn't start")
        fr_instance.cleanup()
        return _json({"stream": False, "message": "Failed to start video stream. Check logs for details."})
    
    return _json({"stream": True, "message": "Success!"})


@app.route("/api/start_fr", methods=["POST"])
def start_fr():
    """For frontend to load FR embeddings and start inference"""
    payload = request.get_json(silent=True)
    data_file = payload.get("data_file")

    try:
        fr_instance.load_embeddings(data_file)
    except (ValueError, FileNotFoundError) as e:
        fr_instance.cleanup()
        log_info(f"Couldn't load embeddings, {e}")
        return _json({"inference": False, "message": "Failed to load embeddings. Check logs for details."})

    fr_instance.start_inference()
    return _json({"inference": True, "message": "Success!"})


@app.route("/api/end", methods=["POST"])
def end():
    """API for frontend to end FR"""

    if not videoplayer.is_started():
        return _json({"stream": False, "message": "Stream not started!"})

    fr_instance.cleanup()
    return _json({"stream": True, "message": "Success!"})


@app.route("/api/status", methods=["GET"])
def status():
    """API to check stream/inference status and last error"""

    return _json({
        "stream_state": videoplayer.stream_state,
        "last_error": videoplayer.last_error,
        "embeddings_loaded": getattr(fr_instance, "embeddings_loaded", False),
        "embeddings_loading": getattr(fr_instance, "embeddings_loading", False),
    })

@app.route("/api/vidFeed", methods=["GET"])
def video_feed():
    return Response(fr_instance.start_video_broadcast(), mimetype="multipart/x-mixed-replace; boundary=frame")


@app.route("/api/frResults", methods=["GET"])
def fr_results():
    return Response(fr_instance.start_detection_broadcast(), mimetype="application/json")


@app.route("/api/submit_settings", methods=["POST"])
def submit_settings():
    try:
        s = fr_instance.fr_settings
        new = {
            "threshold": float(request.form.get("threshold", s["threshold"])),
            "holding_time": float(request.form.get("holding_time", s["holding_time"])),
            "max_detections": int(request.form.get("max_detections", s["max_detections"])),
            "perf_logging": "perf_logging" in request.form,
            "frame_skip": int(request.form.get("frame_skip", s["frame_skip"])),
            "max_broadcast_fps": int(request.form.get("max_broadcast_fps", s["max_broadcast_fps"])),

            "use_differentiator": bool(request.form.get("use_differentiator", None) == "on"),
            "threshold_lenient_diff": float(request.form.get("threshold_lenient_diff", s["threshold_lenient_diff"])),
            "similarity_gap": float(request.form.get("similarity_gap", s["similarity_gap"])),
            
            "use_persistor": bool(request.form.get("use_persistor", None) == "on"),
            "q_max_size": int(request.form.get("q_max_size", s["q_max_size"])),
            "threshold_iou": float(request.form.get("threshold_iou", s["threshold_iou"])),
            "threshold_sim": float(request.form.get("threshold_sim", s["threshold_sim"])),
            "threshold_lenient_pers": float(request.form.get("threshold_lenient_pers", s["threshold_lenient_pers"])),

            "use_low_light_enhancement": bool(request.form.get("use_low_light_enhancement", None) == "on"),
            "use_dehaze": bool(request.form.get("use_dehaze", None) == "on"),
            "use_denoise": bool(request.form.get("use_denoise", None) == "on"),
            "use_lafs": bool(request.form.get("use_lafs", None) == "on"),
        }

        fr_instance.adjust_values(new)
        return _json({"message":"success"}, 200)
    except:
        return _json({"message":"Failed to update settings"}, 500)


@app.route("/api/get_settings", methods=["GET"])
def get_settings():
    return _json(fr_instance.fr_settings)


@app.route("/api/listCameras", methods=["GET"])
def list_cameras():
    """API to list available camera devices"""
    response = _json(list_camera_devices())
    response.headers["Cache-Control"] = "no-store"
    return response


@app.route("/api/get_performance", methods=["GET"])
def get_performance():
    """API to get backend performance"""
    if fr_instance.last_perf is not None:
        return _json(fr_instance.last_perf)
    return _json({})


BENCHMARK_UPLOAD_DIR = os.path.join("data", "benchmark_uploads")
os.makedirs(BENCHMARK_UPLOAD_DIR, exist_ok=True)


@app.route("/api/benchmark/upload", methods=["POST"])
def benchmark_upload():
    """API to upload a video file for offline benchmark analysis"""
    file = request.files.get("video")
    if not file or not file.filename:
        return _json({"message": "No video file provided"}, 400)

    # Clean up any previous upload(s) before saving the new one, so the folder never accumulates
    # files beyond the one currently needed for this session's processing + replay
    for old_file in os.listdir(BENCHMARK_UPLOAD_DIR):
        try:
            os.remove(os.path.join(BENCHMARK_UPLOAD_DIR, old_file))
        except OSError:
            pass

    filename = f"{uuid.uuid4().hex}_{file.filename}"
    path = os.path.join(BENCHMARK_UPLOAD_DIR, filename)
    file.save(path)
    return _json({"video_path": path, "video_url": f"/data/benchmark_uploads/{filename}"})


@app.route("/api/benchmark/run", methods=["GET"])
def benchmark_run():
    """API to run the offline video benchmark, streaming per-frame progress"""
    video_path = request.args.get("video_path")
    if not video_path:
        return _json({"message": "video_path is required"}, 400)

    # Restrict to the upload directory - video_path is client-supplied, so this prevents
    # it from being used to open arbitrary files elsewhere on disk
    resolved = os.path.realpath(video_path)
    upload_dir = os.path.realpath(BENCHMARK_UPLOAD_DIR)
    if not resolved.startswith(upload_dir + os.sep) or not os.path.exists(resolved):
        return _json({"message": "Invalid video_path"}, 404)

    return Response(fr_instance.run_video_benchmark(resolved), mimetype="application/json")


@app.route("/api/benchmark/export", methods=["POST"])
def benchmark_export():
    """
    API to export a video benchmark's run history as an Excel report - one row per run,
    so the same video benchmarked repeatedly under different model settings can be compared
    side-by-side (models used vs. resulting confidence scores).
    """
    data = request.get_json(silent=True) or {}
    runs = data.get("runs") or []
    thumbnail_b64 = data.get("thumbnail")

    wb = Workbook()
    ws = wb.active
    ws.title = "Benchmark Report"

    for col, width in {1: 22, 2: 28, 3: 30, 4: 16, 5: 14, 6: 14, 7: 14, 8: 16}.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    row = 1
    if thumbnail_b64:
        try:
            image_bytes = base64.b64decode(thumbnail_b64.split(",")[-1])
            img = XLImage(BytesIO(image_bytes))
            if img.width > 480:
                scale = 480 / img.width
                img.width = 480
                img.height = int(img.height * scale)
            ws.add_image(img, "A1")
            row = int(img.height / 15) + 2  # approximate rows spanned by the image, plus spacing
        except Exception as e:
            log_info(f"Failed to embed benchmark thumbnail: {e}")

    bold = Font(bold=True)

    ws.cell(row=row, column=1, value="Report generated").font = bold
    ws.cell(row=row, column=2, value=datetime.now().isoformat(timespec="seconds"))
    row += 1
    ws.cell(row=row, column=1, value="Recognition threshold").font = bold
    ws.cell(row=row, column=2, value=fr_instance.fr_settings.get("threshold"))
    row += 2

    headers = [
        "Run time", "Models used", "People seen", "Detections captured",
        "Mean distance", "Min distance (best)", "Max distance (worst)", "% identified",
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=row, column=col, value=header).font = bold
    row += 1

    for run in runs:
        stats = run.get("stats") or {}
        toggles = run.get("toggles") or []
        people_seen = run.get("people_seen") or []
        values = [
            run.get("timestamp", ""),
            ", ".join(toggles) if toggles else "none",
            ", ".join(people_seen) if people_seen else "none",
            stats.get("detections_captured"),
            stats.get("mean"),
            stats.get("min"),
            stats.get("max"),
            stats.get("pct_identified"),
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row=row, column=col, value=value)
        row += 1

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"benchmark_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        buffer,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )
# ───────────────────────────── Pages ──────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/old_layout")
def old_layout():
    return render_template("old_layout.html")

@app.route("/welcome")
def welcome():
    return render_template("welcome.html")

@app.route("/seats")
def seats():
    return render_template("seats.html")

@app.route("/settings")
def settings():
    return render_template("settings.html", **fr_instance.fr_settings) # Render with fr_settings as context

@app.route("/benchmark")
def benchmark():
    return render_template("benchmark.html")

@app.route("/benchmark-history")
def benchmark_history_page():
    return render_template("benchmark_history.html")

@app.route('/data/<path:filename>')
def serve_data(filename):
    return send_from_directory('data', filename)


# ───────────────────────────── Helpers ────────────────────────────────────

def _json(data, status: int | None = 200, headers: dict | None = None) -> Response:
    '''Wrapper around json.dumps. Takes in data (Python list/dict) and optional status and headers'''
    response = Response(json.dumps(data), status=status, mimetype='application/json')
    if headers:
        for key, value in headers.items():
            response.headers[key] = value
    return response


def _collect_reference_images() -> dict:
    images = {}

    # From captures
    captures = os.path.join("data", "captures")
    if os.path.isdir(captures):
        for name in os.listdir(captures):
            person = os.path.join(captures, name)
            if os.path.isdir(person):
                imgs = [f"/data/captures/{name}/{i}" for i in os.listdir(person) if i.lower().endswith((".jpg", ".jpeg", ".png"))]
                if imgs:
                    images[name] = images.get(name, []) + imgs

    return images


# ───────────────────────────── Main ───────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Facial Recognition")
    parser.add_argument("-ip", "--ipaddress", type=str)
    parser.add_argument("-p", "--port", type=int)
    parser.add_argument("--env", type=str, choices=["development", "production"])
    parser.add_argument("--prod", action="store_true")
    args = parser.parse_args()

    if args.ipaddress:
        config.ip = args.ipaddress
    if args.port:
        config.port = args.port
    if args.env:
        config.env = args.env
    if args.prod:
        config.env = "production"

    # Cleanup upon interruption
    signal.signal(signal.SIGINT, lambda sig, frame: fr_instance.cleanup(force_exit=True))

    if config.env == "production":
        try:
            log_info(f"Production mode on {config.ip}:{config.port}")
            serve(app, host=config.ip, port=config.port)
        except Exception as e:
            log_info(f"Waitress failed ({e}), using Flask")
            app.run(host=config.ip, port=config.port, use_reloader=False)
    else:
        log_info(f"Development mode on {config.ip}:{config.port}")
        app.run(debug=True, host=config.ip, port=config.port, use_reloader=False)
